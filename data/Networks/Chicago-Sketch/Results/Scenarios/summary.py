import os
import openpyxl as xl


global wb,sheet,city,nScens,nInstances




def openWb(name):
	'''
	Description
	Args:
		arg1(type):description.
	Return:
		Return(type):description.
	'''
	global wb	
	try:
		wb=xl.load_workbook(f'{name}.xlsx')
	except:
		wb=xl.Workbook()
		wb.save(f'{name}.xlsx')
		
def modSheet(pSheet):
	'''
	ROUTINE:
		modSheet(pSheet)
	PURPOSE:
		Cambia la hoja activa a la hoja pasada por parámetro. Si no hay ningúna hoja con dicho nombre la crea.

	ARGUMENTS:
		pSheet(string): El nombre de la hoja que se quiere activar
	
	RETURN VALUE:
		None

	EXAMPLE:		
		modSheet('Hoja1')
	'''	
	global sheet
	try:
		sheet=wb[pSheet]
		return sheet
	except:
		wb.create_sheet(pSheet)
		sheet=wb[pSheet]
		return sheet

def analyze(nPhases):
	'''
	Description
	Args:
		arg1(type):description.
	Return:
		Return(type):description.
	'''

	openWb('summary')
	cTimes = modSheet(f'compTimes{nPhases}')
	cost = modSheet(f'objFun{nPhases}')
	rel = modSheet(f'reliability{nPhases}')
	
	
	scens=list(range(1,nScens+1))+['Total']
	files={s:open(f'PHFit{nPhases}/PHFit{nPhases}_scen{s}.txt') for s in scens}

	insts=readInstances()
	r=3
	for s,t in insts.keys():
		col=2		
		for sc in scens:
			l=files[sc].readline().replace('\n','').split('\t')
			nums=list(map(float,l[:-1]))			
			
			if l!=['']:
				if nums[0]==s and nums[1]==t:					
					cTimes.cell(r,col).value=nums[2]
					cost.cell(r,col).value=nums[3]
					rel.cell(r,col).value=nums[5]
			col+=1
								
		r+=1
	wb.save(f'summary.xlsx')

def readInstances():
	'''
	Description
	Args:
		arg1(type):description.
	Return:
		Return(type):description.
	'''
	f=open(f'../../Chicago-Sketch_Instances.txt')
	insts={}
	for l in f:
		if l[0]!='#':
			n=list(map(float,l.replace('\n','').split('\t')))
			insts[n[0],n[1]]=n[2]
	f.close()
	return insts


if __name__ == '__main__':
	city='Chicago-Sketch'
	nScens=5
	nInstances=40
	nPhases=3
	
	
	analyze(nPhases)
